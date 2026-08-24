#!/usr/bin/env python3
"""
Read data/founders.xlsx and data/timeline.xlsx back into the JSON the website
is built from. Run this after editing either spreadsheet, then run build.sh.

Nothing is overwritten unless the spreadsheet parses cleanly, and the previous
JSON is kept alongside as <name>.json.bak.
"""
import json, os, shutil, sys
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, "data")
CATS = {"founding", "expansion", "publications", "members", "conventions", "insignia"}


def rows_of(path, sheet=0):
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[sheet]
    header_row = None
    for r in range(1, 12):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 4)]
        if vals and str(vals[0]).strip() in ("id", "year"):
            header_row = r
            break
    if header_row is None:
        sys.exit(f"{os.path.basename(path)}: could not find the header row.")
    heads = []
    c = 1
    while True:
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            break
        heads.append(str(v).strip())
        c += 1
    out = []
    for r in range(header_row + 1, ws.max_row + 1):
        row = {h: ws.cell(row=r, column=i + 1).value for i, h in enumerate(heads)}
        filled = [v for v in row.values() if v is not None and str(v).strip()]
        if len(filled) < 2:      # blank rows, and the "add below this line" note
            continue
        out.append(row)
    return out


def text(v):
    return str(v).strip() if v is not None and str(v).strip() else ""


def year(v):
    s = text(v)
    if not s:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def lines(v):
    return [l.strip() for l in text(v).replace("\r", "").split("\n") if l.strip()]


def save(name, obj):
    fp = os.path.join(DATA, name)
    if os.path.exists(fp):
        shutil.copyfile(fp, fp + ".bak")
    json.dump(obj, open(fp, "w"), indent=1)


# ----------------------------------------------------------------- founders
fp = os.path.join(DATA, "founders.xlsx")
if os.path.exists(fp):
    out, problems = [], []
    for n, r in enumerate(rows_of(fp), start=1):
        fid, nm = text(r.get("id")), text(r.get("name"))
        if not fid or not nm:
            problems.append(f"row {n}: needs both an id and a name")
            continue
        links = []
        for l in lines(r.get("links")):
            if "|" in l:
                lab, url = l.split("|", 1)
                links.append({"label": lab.strip(), "url": url.strip(), "verified": True})
            elif l.startswith("http"):
                links.append({"label": "Read more", "url": l, "verified": True})
            else:
                problems.append(f"{nm}: link “{l[:40]}” should read  Label | https://address")
        # "claim ~ what the sources say ~ how to settle it", one per line
        uncs = []
        for l in lines(r.get("uncertainties")):
            parts = [x.strip() for x in l.split("~")]
            if len(parts) == 1:
                problems.append(f"{nm}: open question “{l[:38]}” needs the three parts "
                                "separated by ~  (claim ~ detail ~ how to settle it)")
            uncs.append({"claim": parts[0],
                         "detail": parts[1] if len(parts) > 1 else "",
                         "how_to_resolve": parts[2] if len(parts) > 2 else ""})
        # "volume-id | page | label", one per line
        srcs = []
        for l in lines(r.get("sources")):
            parts = [x.strip() for x in l.split("|")]
            if not parts[0]:
                continue
            pg = year(parts[1]) if len(parts) > 1 else None
            srcs.append({"doc": parts[0], "page": pg,
                         "label": parts[2] if len(parts) > 2 else parts[0]})
        out.append({
            "id": fid, "name": nm, "honorific": text(r.get("honorific")), "sort": n,
            "class_year": year(r.get("class_year")) or 1836,
            "role": text(r.get("role")), "profession": text(r.get("profession")),
            "born": {"year": year(r.get("born_year")), "date": text(r.get("born_date")),
                     "place": text(r.get("born_place"))},
            "died": {"year": year(r.get("died_year")), "date": text(r.get("died_date")),
                     "place": text(r.get("died_place"))},
            "buried": {"place": text(r.get("buried_place")),
                       "findagrave": text(r.get("findagrave"))},
            "family": text(r.get("family")), "bio": text(r.get("bio")),
            "achievements": lines(r.get("achievements")),
            "annals": text(r.get("annals")),
            "portrait": text(r.get("portrait")) or None,
            "links": links, "sources": srcs, "uncertainties": uncs,
        })
    for p in problems:
        print("  ! " + p)
    if out:
        save("founders.json", out)
        filled = sum(1 for f in out if f["bio"])
        graves = sum(1 for f in out if f["buried"]["findagrave"])
        opens = sum(len(f["uncertainties"]) for f in out)
        print(f"  founders.json: {len(out)} founders, {filled} with a written biography, "
              f"{graves} with a Find a Grave link, {opens} open questions still listed")

# ----------------------------------------------------------------- timeline
fp = os.path.join(DATA, "timeline.xlsx")
if os.path.exists(fp):
    out, problems = [], []
    for n, r in enumerate(rows_of(fp), start=1):
        y, title = year(r.get("year")), text(r.get("title"))
        if not y or not title:
            problems.append(f"row {n}: needs both a year and a title")
            continue
        cat = text(r.get("category")).lower() or "founding"
        if cat not in CATS:
            problems.append(f"{y} “{title[:30]}”: category “{cat}” is not one of "
                            + ", ".join(sorted(CATS)))
            cat = "founding"
        e = {"year": y, "category": cat, "title": title, "text": text(r.get("text"))}
        if text(r.get("date")):
            e["date"] = text(r.get("date"))
        if text(r.get("feature")).lower() in ("yes", "y", "true", "1"):
            e["feature"] = True
        if text(r.get("link")):
            e["link"] = text(r.get("link"))
            e["link_label"] = text(r.get("link_label")) or "Open"
        if text(r.get("source_doc")) and year(r.get("source_page")):
            e["source"] = {"doc": text(r.get("source_doc")),
                           "page": year(r.get("source_page")),
                           "label": text(r.get("source_label")) or "Source"}
        if text(r.get("uncertain")):
            e["uncertain"] = text(r.get("uncertain"))
        out.append(e)
    for p in problems:
        print("  ! " + p)
    if out:
        save("timeline-core.json", out)
        unc = sum(1 for e in out if e.get("uncertain"))
        print(f"  timeline-core.json: {len(out)} hand-written events, {unc} marked not settled "
              f"(chapter charterings and closures are added by build_timeline.py)")

print("  now run ./build/build.sh")
