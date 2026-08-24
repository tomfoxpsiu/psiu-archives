#!/usr/bin/env python3
"""
Write data/founders.xlsx and data/timeline.xlsx from the current JSON, so the
content can be edited in a spreadsheet instead of a text file.

    ./build/make_sheets.py        make (or refresh) the spreadsheets
    ./build/import_sheets.py      read them back into the JSON

Refreshing overwrites the spreadsheets, so import before you make.
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, "data")

ARIAL   = "Arial"
HEAD_F  = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
HEAD_BG = PatternFill("solid", fgColor="7A1428")
EDIT_BG = PatternFill("solid", fgColor="FFF7D6")     # type here
AUTO_BG = PatternFill("solid", fgColor="EFEFEF")     # do not edit
BODY_F  = Font(name=ARIAL, size=10)
NOTE_F  = Font(name=ARIAL, size=9, italic=True, color="666666")
TITLE_F = Font(name=ARIAL, size=13, bold=True, color="7A1428")
THIN    = Side(style="thin", color="D9D9D9")
BORDER  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def sheet_header(ws, title, legend, row=1):
    ws.cell(row=row, column=1, value=title).font = TITLE_F
    ws.cell(row=row + 1, column=1, value=legend).font = NOTE_F
    return row + 3


def write_table(ws, start, columns, rows, locked=()):
    """columns: [(header, width, note)] ; rows: list of dicts keyed by header."""
    for c, (head, width, note) in enumerate(columns, start=1):
        cell = ws.cell(row=start, column=c, value=head)
        cell.font, cell.fill, cell.border = HEAD_F, HEAD_BG, BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if note:
            cell.comment = Comment(note, "Psi Upsilon Digital Museum")
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[start].height = 30
    ws.freeze_panes = ws.cell(row=start + 1, column=1)

    for r, data in enumerate(rows, start=start + 1):
        for c, (head, _w, _n) in enumerate(columns, start=1):
            cell = ws.cell(row=r, column=c, value=data.get(head, ""))
            cell.font, cell.border = BODY_F, BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.fill = AUTO_BG if head in locked else EDIT_BG
        ws.row_dimensions[r].height = 46
    return start + 1 + len(rows)


# ----------------------------------------------------------------- founders
def founders_sheet():
    founders = json.load(open(os.path.join(DATA, "founders.json")))
    cols = [
        ("id", 26, "Do not change — this is how the page is named."),
        ("name", 24, "The full name as it should appear as a page heading."),
        ("honorific", 12, "The Rev. Dr., The Hon., Colonel, Dr. — or leave blank."),
        ("class_year", 10, "1836 or 1837."),
        ("role", 30, "One short line shown as a badge under the name."),
        ("profession", 26, "How he made his living."),
        ("born_year", 10, "A four-digit year, or leave blank."),
        ("born_place", 30, "Town, county, state."),
        ("died_year", 10, "A four-digit year."),
        ("died_place", 26, "Where he died."),
        ("buried_place", 30, "Cemetery, town, state."),
        ("findagrave", 40, "Paste the whole address of his Find a Grave memorial page."),
        ("family", 46, "Parents, wife, children — written as a sentence or two."),
        ("bio", 72, "The Fraternity's own account of his life. Leave a blank line between "
                    "paragraphs (Alt+Enter inside the cell)."),
        ("achievements", 72, "One per line inside the cell (Alt+Enter between them)."),
        ("annals", 72, "The 1941 Annals passage. Already filled in — change only to correct it."),
        ("portrait", 34, "A file name you have put in site/assets/people, e.g. "
                         "assets/people/edward-martindale.jpg"),
        ("links", 46, "One per line as: Label | https://address"),
    ]
    rows = []
    for f in sorted(founders, key=lambda x: x.get("sort", 99)):
        rows.append({
            "id": f["id"], "name": f["name"], "honorific": f.get("honorific", ""),
            "class_year": f["class_year"], "role": f.get("role", ""),
            "profession": f.get("profession", ""),
            "born_year": f["born"].get("year") or "", "born_place": f["born"].get("place", ""),
            "died_year": f["died"].get("year") or "", "died_place": f["died"].get("place", ""),
            "buried_place": f["buried"].get("place", ""),
            "findagrave": f["buried"].get("findagrave", ""),
            "family": f.get("family", ""), "bio": f.get("bio", ""),
            "achievements": "\n".join(f.get("achievements") or []),
            "annals": f.get("annals", ""), "portrait": f.get("portrait") or "",
            "links": "\n".join("%s | %s" % (l.get("label", "Read more"), l["url"])
                               for l in (f.get("links") or [])),
        })
    wb = Workbook()
    ws = wb.active
    ws.title = "Founders"
    start = sheet_header(
        ws, "The seven founders",
        "Type into the pale yellow cells. Grey cells are used to name the page — leave them "
        "alone. Blank cells show as “not yet recorded” on the website, which is fine. "
        "Save the file, then run ./build/import_sheets.py and ./build/build.sh.")
    write_table(ws, start, cols, rows, locked=("id",))
    wb.save(os.path.join(DATA, "founders.xlsx"))
    return len(rows)


# ----------------------------------------------------------------- timeline
def timeline_sheet():
    core = json.load(open(os.path.join(DATA, "timeline-core.json")))
    cols = [
        ("year", 8, "The year. Required."),
        ("date", 20, "A fuller date if you have one, e.g. 24 November 1833. Optional."),
        ("category", 14, "One of: founding, expansion, publications, members, conventions, "
                         "insignia."),
        ("title", 40, "A short headline, e.g. “The first Convention”."),
        ("text", 82, "One or two sentences."),
        ("feature", 10, "Put yes to give the entry a highlighted box. Otherwise leave blank."),
        ("link", 40, "A page on the museum site, e.g. people/william-howard-taft.html. Optional."),
        ("link_label", 24, "The words for that link, e.g. “William Howard Taft”."),
        ("source_doc", 40, "The id of the volume this came from, if any — the last part of its "
                           "web address."),
        ("source_page", 12, "The page number in that volume."),
        ("source_label", 30, "How to name the source, e.g. “Annals of Psi Upsilon, Part 2”."),
    ]
    rows = []
    for e in core:
        src = e.get("source") or {}
        rows.append({
            "year": e["year"], "date": e.get("date", ""), "category": e.get("category", ""),
            "title": e["title"], "text": e.get("text", ""),
            "feature": "yes" if e.get("feature") else "",
            "link": e.get("link", ""), "link_label": e.get("link_label", ""),
            "source_doc": src.get("doc", ""), "source_page": src.get("page", ""),
            "source_label": src.get("label", ""),
        })
    wb = Workbook()
    ws = wb.active
    ws.title = "Timeline"
    start = sheet_header(
        ws, "Timeline events",
        "Add a row for each new event; the order does not matter, the website sorts by year. "
        "Chapter charterings and closures are added automatically from the chapter roll, so do "
        "not list them here. Save, then run ./build/import_sheets.py and ./build/build.sh.")
    end = write_table(ws, start, cols, rows)
    ws.cell(row=end + 1, column=1, value="Add new events below this line.").font = NOTE_F
    wb.save(os.path.join(DATA, "timeline.xlsx"))
    return len(rows)


if __name__ == "__main__":
    n = founders_sheet()
    m = timeline_sheet()
    print(f"  data/founders.xlsx  ({n} founders)")
    print(f"  data/timeline.xlsx  ({m} events; chapter events stay automatic)")
